import time
import openpyxl
import json
import pyproj
import os
import warnings
import requests

query = {
    "site": "Site",
    "ui": 3,
    "network": "sernhac_MC1_GW3",
    "engine": "2",
    "transmitter": {
        "lat": "38.990511",
        "lon": "1.447991",
        "alt": "0.2",
        "frq": "2400",
        "txw": 0.00100145,
        "bwi": "0.1"
    },
    "receiver": {
        "lat": 0,
        "lon": 0,
        "alt": "1",
        "rxg": "2",
        "rxs": "-103"
    },
    "antenna": {
        "txg": "1.8",
        "txl": "0.1",
        "ant": "39",
        "azi": "165",
        "tlt": "33",
        "hbw": "0",
        "vbw": "0",
        "fbr": "1.8",
        "pol": "v"
    },
    "model": {
        "pm": "2",
        "pe": "2",
        "ked": "0",
        "rel": "95"
    },
    "environment": {
        "clm": "1",
        "cll": "2",
        "clt": "Minimal.clt"
    },
    "output": {
        "units": "m",
        "col": "RAINBOW.dBm",
        "out": "2",
        "ber": None,
        "mod": None,
        "nf": "-120",
        "res": "30",
        "rad": "20"
    }
}

#permet de onvertir des coordonnées EPSG en GPS
def transform(x,y) -> list:
    inProj = pyproj.Proj(init='epsg:2154')
    outProj = pyproj.Proj(init='epsg:4326')
    x2, y2 = pyproj.transform(inProj, outProj, x, y)
    return [x2, y2]

#retourne le nom de la ville ou se trouve la centrale
def get_site_name(path) -> str:
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    for i in range(sheet_obj.max_row):
        cell_obj = sheet_obj.cell(row=i + 1, column=1)
        if cell_obj.value == "solar_park_name":
            return sheet_obj.cell(row=i + 1, column=2).value

#vérifier que la chaine correspond bien à un réseau
def isNetwork(string) :
    if string is not None:
        mc = str(string).find("MC")
        gw = str(string).find("GW")
        return gw != -1 and mc != -1

#liste des réseaux et la coordonnée de leur GW
def post_network_GW(path):
    warnings.filterwarnings("ignore") #pour éviter de surcharger la console avec les warnings
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    type = get_type(path)
    for i in range(sheet_obj.max_row):
        cell_obj = sheet_obj.cell(row=i + 1, column=1)
        if isNetwork(cell_obj.value):
            if type != 4326:
                transf = transform(float(str(sheet_obj.cell(row=i + 1, column=5).value).split(',')[0]),float(str(sheet_obj.cell(row=i + 1, column=5).value).split(',')[1]))
            else:
                transf = sheet_obj.cell(row=i + 1, column=5).value.split(',')
                transf = [transf[1],transf[0]]
            query.update({"site": "GW"})
            query.update({"network": get_site_name(path) + "." + str(cell_obj.value).replace('GW','').replace('MC','').replace('_','.')}),
            query.update(
                {"transmitter" : {
                    "lon": transf[0],
                    "lat" : transf[1],
                    "alt": "0.2",
                    "frq": "2400",
                    "txw": 0.00100145,
                    "bwi": "0.1"
                    }
                })
            print("posting " + get_site_name(path) + '_' + str(cell_obj.value) + "...")
            try:
                os.mkdir("generation")
            except FileExistsError:
                pass
            with open('generation/geojson' + get_site_name(path) + str(cell_obj.value) + '.json', 'w') as mon_fichier:
                json.dump(query, mon_fichier, indent=4)
            print(post(query))
            print(cell_obj.value + " posted !")
            time.sleep(1)

#poster les information stockées dans le fichier json
def post(json):
    headers = {
        "key"  : "47474-8f9ffce24b48f4dca92bfe201421a5e3f601989d",
        "content-Type" : "application/json, text/javascript, */*; q=0.01"
    }
    req = requests.post(url = "https://api.cloudrf.com/area",json= json,headers=headers)
    return str(req.content).replace('\\n','')

#générer le fichier CSV contenant les coordonnées de tous les TC de la centrale
def generate_csv_all(path):
    print("generating csv file...")
    ligneEntete = "id,longitude,latitude\n"
    valeurs = []
    count = 0
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    type = get_type(path)
    for i in range(sheet_obj.max_row):
        cell_obj = sheet_obj.cell(row=i + 1, column=10)
        if cell_obj.value != "TC coordinates" and cell_obj.value is not None:
            if type != 4326:
                coord = transform(float(cell_obj.value.split(',')[0]),
                                  float(cell_obj.value.split(',')[1]))
                coord = [coord[1],coord[0]]
            else:
                coord = str(cell_obj.value).split(',')
            valeurs.append(str(sheet_obj.cell(row=i + 1, column=3).value) + "," + str(coord[1]) + "," + str(coord[0]) + "\n")
    f = open('tc_coordinates.csv', 'w')
    f.write(ligneEntete)
    for valeur in valeurs:
        f.write(valeur)

#génère un fihcier csv par réseau
def generate_csv_all_nt(path):
    for nt in get_all_nt(path):
        generate_csv_nt(path,nt)

#génère le fichier csv du réseau en paramètres
def generate_csv_nt(path,nt):
    print("generating csv file for " + nt + "...")
    ligneEntete = "id,longitude,latitude\n"
    valeurs = []
    count = 0
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    type = get_type(path)
    for i in range(sheet_obj.max_row):
        cell_obj = sheet_obj.cell(row=i + 1, column=10)
        actnt = str(sheet_obj.cell(row = i + 1, column = 1).value) + "." + str(sheet_obj.cell(row = i + 1, column =  2).value)
        if cell_obj.value != "TC coordinates" and cell_obj.value is not None and actnt == nt:
            if type != 4326:
                coord = transform(float(cell_obj.value.split(',')[0]),
                                  float(cell_obj.value.split(',')[1]))
                coord = [coord[1],coord[0]]
            else:
                coord = str(cell_obj.value).split(',')
            valeurs.append(str(sheet_obj.cell(row=i + 1, column=3).value) + "," + str(coord[1]) + "," + str(coord[0]) + "\n")
    f = open('tc_coordinates' + nt + '.csv', 'w')
    f.write(ligneEntete)
    for valeur in valeurs:
        f.write(valeur)
#supprimer un élément de la base (par rapport à son ID)
def delete(elt):
    headers = {
        "key": "47474-8f9ffce24b48f4dca92bfe201421a5e3f601989d",
        "content-Type": "application/json, text/javascript, */*; q=0.01"
    }
    print("deleting " + str(elt) + "...")
    req = requests.get(url="https://api.cloudrf.com/archive/delete?cid=" + str(elt), headers=headers)
    print("code : " + str(req.status_code) + " message : " + str(req._content))

#nettoyer une réponse HTTP
def nettoyer(list,string):
    for el in list:
        string.replace(el,'')
    return string

#returne le type de coordonnées du fichier config (GPS ou EPSG)
def get_type(path) -> str:
    warnings.filterwarnings("ignore")
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    for i in range(sheet_obj.max_row):
        cell_obj = sheet_obj.cell(row=i + 1, column=1).value
        if cell_obj == 'coordinates_system':
            return sheet_obj.cell(row=i + 1, column=2).value

#renvoie la liste de tous les TC présents sur le ficher config dans un dictionnaire
def get_TC_list(path) -> dict:
    l = dict()
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    type = get_type(path)
    for i in range(sheet_obj.max_row):
        cell_obj = sheet_obj.cell(row=i + 1, column=10)
        if cell_obj.value != "TC coordinates" and cell_obj.value is not None:
            if type != 4326:
                transf = transform(float(str(sheet_obj.cell(row=i + 1, column=10).value).split(',')[0]),float(str(sheet_obj.cell(row=i + 1, column=10).value).split(',')[1]))
            else:
                transf = sheet_obj.cell(row=i + 1, column=10).value.split(',')
                transf = [transf[1], transf[0]]
            MC = sheet_obj.cell(row=i + 1, column=1).value
            GW = sheet_obj.cell(row=i + 1, column=2).value
            ID = sheet_obj.cell(row=i + 1, column=3).value
            lon = transf[0]
            lat = transf[1]
            l.update({str(i) :
            {
                "MC" : MC,
                "GW" : GW,
                "ID" : ID,
                "lon" : lon,
                "lat" : lat,
            }})
            print(l)
    return l

#permet de faire un POST avec l'url en paramètres
def post_on_url(url):
    headers = {
        "key": "47474-8f9ffce24b48f4dca92bfe201421a5e3f601989d",
        "content-Type": "application/json, text/javascript, */*; q=0.01"
    }
    req = requests.post(url=url, headers=headers)
    print("code : " + str(req.status_code) + " message : " + str(req._content))

#permet de faire un GET avec l'url en paramètres
def get_on_url(url):
    headers = {
        "key": "47474-8f9ffce24b48f4dca92bfe201421a5e3f601989d",
        "content-Type": "application/json, text/javascript, */*; q=0.01"
    }
    print("calculation in progress ...")
    req = requests.get(url=url, headers=headers)
    response = req.content
    response1 = str(response).replace('\\n', '')
    return response1

#retourne la liste des networks présents dans le fichier config
def get_all_nt(path) -> list:
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    type = get_type(path)
    l = list()
    for i in range(sheet_obj.max_row):
        cell_obj = sheet_obj.cell(row=i + 1, column=1)
        if isNetwork(cell_obj.value):
            l.append(str(cell_obj.value).replace('GW','').replace('MC','').replace('_','.'))
    return l

#supprime tous les Networks
def delete_all_Nt(path):
    l = get_all_nt(path)
    site_name = get_site_name(path)
    for nt in l:
        get_on_url("://api.cloudrf.com/archive/delete/network?nid=" + site_name + "." + nt)

#merge tous les réseaux
def merge_all(path):
    l = get_all_nt(path)
    site_name = get_site_name(path)
    for nt in l:
        print("merging" + site_name + "." + nt)
        get_on_url("https://api.cloudrf.com/mesh?network=" + site_name + "." + nt + "&name=" + nt + "_merged" )
        print(site_name + "." + nt + "merged !")

#envoie tous les TC sur la base (réseaux créés automatiquement grâce aux noms)
def post_all_TC(path):
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    type = get_type(path)
    for i in range(sheet_obj.max_row):
        cell_obj = sheet_obj.cell(row=i + 1, column=10)
        if cell_obj.value != "TC coordinates" and cell_obj.value is not None:
            if type != 4326:
                transf = transform(float(str(sheet_obj.cell(row=i + 1, column=10).value).split(',')[0]),float(str(sheet_obj.cell(row=i + 1, column=10).value).split(',')[1]))
            else:
                transf = sheet_obj.cell(row=i + 1, column=10).value.split(',')
                transf = [transf[1], transf[0]]
            MC = sheet_obj.cell(row=i + 1, column=1).value
            GW = sheet_obj.cell(row=i + 1, column=2).value
            ID = sheet_obj.cell(row=i + 1, column=3).value
            lon = transf[0]
            lat = transf[1]
            site = get_site_name(path)
            print(post_TC(MC,GW,ID,lat,lon,site))
    merge_NT(str(site) + "." + str(MC) + "." + str(GW))
            #time.sleep(0.5)

#envoyer un TC sur la base avec toutes les infos qui vont avec
def post_TC(MC, GW, TC_id, lat, long, site) -> str:
    print("posting TC " + str(MC) + "." + str(GW) + "." + str(TC_id) + " ...")
    TC = {
        "site": str(TC_id),
        "ui": 3,
        "network": str(site) + "." + str(MC) + "." + str(GW),
        "engine": "2",
        "transmitter": {
            "lat": str(lat),
            "lon": str(long),
            "alt": "0.2",
            "frq": "2400",
            "txw": 0.00100145,
            "bwi": "0.1"
        },
        "receiver": {
            "lat": 0,
            "lon": 0,
            "alt": "1",
            "rxg": "2",
            "rxs": "-103"
        },
        "antenna": {
            "txg": "1.8",
            "txl": "0.1",
            "ant": "39",
            "azi": "165",
            "tlt": "33",
            "hbw": "0",
            "vbw": "0",
            "fbr": "1.8",
            "pol": "v"
        },
        "model": {
            "pm": "2",
            "pe": "2",
            "ked": "0",
            "rel": "95"
        },
        "environment": {
            "clm": "1",
            "cll": "2",
            "clt": "Minimal.clt"
        },
        "output": {
            "units": "m",
            "col": "RAINBOW.dBm",
            "out": "2",
            "ber": None,
            "mod": None,
            "nf": "-120",
            "res": "30",
            "rad": "20"
        }
    }
    print(post(TC))
    return "TC " + str(TC_id) + " posted !"

#retourne un dictionnaire contenant tous les ID des TC
def get_all_id() -> dict:
    headers = {
        "key": "47474-8f9ffce24b48f4dca92bfe201421a5e3f601989d",
        "content-Type": "application/json, text/javascript, */*; q=0.01"
    }
    print("gettting IDs... ")
    req = requests.get(url="https://api.cloudrf.com/archive/data?", headers=headers)
    response  = req.content
    response1 = str(response).replace('\\n','')
    return json.loads(str(response).replace('\\n','')[2: len(response1) - 1])

#va merger le Network en paramètre
def merge_NT(nt):
    headers = {
        "key": "47474-8f9ffce24b48f4dca92bfe201421a5e3f601989d",
        "content-Type": "application/json, text/javascript, */*; q=0.01"
    }
    print("merging " + str(nt) + "...")
    req = requests.get(url="https://api.cloudrf.com/mesh?network=" + str(nt) + "&name=" + str(nt) + "_merged" , headers=headers)
    print("code : " + str(req.status_code) + " message : " + str(req._content))

#supprime tous les éléments de la base
def delete_all():
    ids = get_all_id()['calcs']
    for el in ids:
        delete(el['id'])
    print("nothing else to delete ...")

#########################################
#       toujours poster GW avant TC     #
#########################################
#delete_all()
#post_network_GW("config_sernhac.xlsx")
#delete_all()
#post_network_GW("config_sernhac.xlsx")
#post_all_TC("config_sernhac.xlsx")
#merge_all("config_gargenville.xlsx")
print(get_all_nt("config_gargenville.xlsx"))
generate_csv_nt("config_gargenville.xlsx","1.1")
# c = input("clean database ? (o/n ; 0 to cancel) : ")
# if c == "o":
#     delete_all()
# while(c!=0):
#     path = input("enter the path :")
#     while(c!=5):
#         print("1- post all GW\n"
#               "2- post all TC\n"
#               "3- post TC + GW\n"
#               "4- delete all\n"
#               "5- change path\n"
#               "0- cancel\n")
#         c = input("select an action :")
#         if c== "1":
#             post_network_GW(path)
#         elif c == "2":
#             post_all_TC(path)
#         elif c == "3":
#             post_network_GW(path)
#             post_all_TC(path)
#         elif c == "4":
#             delete_all()
#         elif c == "0":
#             print("bye !")
