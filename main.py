# This is a sample Python script.
import os
import pyproj
import openpyxl
import json
from selenium import webdriver
from time import sleep
import csv
import pandas as pd
# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.


def print_hi(name):
    # Use a breakpoint in the code line below to debug your script.
    print(f'Hi, {name}')  # Press Ctrl+F8 to toggle the breakpoint.

import folium
import pyproj
def tranform(x,y):
    inProj = pyproj.Proj(init='epsg:2154')
    outProj = pyproj.Proj(init='epsg:4326')
    x2, y2 = pyproj.transform(inProj, outProj, x, y)
    return [x2, y2]

# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    # Give the location of the file
    #path = "config_sernhac.xlsx"###################################

    print("==================================================================== APPLICATION =======================================================================")
    path = input("Enter the .xlsx file path : ")
    coord = input("type of coordinate (epsg or gps) :")
    # Create directory
    dirName = 'files'

    # Create target Directory
    try:
        os.mkdir(dirName)
    except FileExistsError:
        pass



    # To open the workbook
    # workbook object is created
    wb_obj = openpyxl.load_workbook(path)

    # Get workbook active sheet object
    # from the active attribute
    sheet_obj = wb_obj.active

    # Cell objects also have a row, column,
    # and coordinate attributes that provide
    # location information for the cell.

    # Note: The first row or
    # column integer is 1, not 0.

    # Cell object is created by using
    # sheet object's cell() method.
    # Création d'un fichier à partir de la liste tableau

    tc_coordinates = open('files/tc_coordinates.csv', 'w')
    print("generating tc_coordinates.csv...")
    for i in range(sheet_obj.max_row):
        cell_obj = sheet_obj.cell(row=i + 1, column=10)
        if (cell_obj.value is not None):
            if(cell_obj.value == 'TC coordinates'):
                tc_coordinates.write('id,longitude,latitude\n')
            else:
                tc_coordinates.write( str(i) + ',' + cell_obj.value + '\n')
    north_coordinates_file = open('files/north_coordinates_file.csv', 'w')
    print("generating north_coordinates_file.csv...")
    north_coordinates_file.write('id,longitude,latitude\n')
    south_coordinates_file = open('files/south_coordinates_file.csv', 'w')
    print("generating south_coordinates_file.csv...")
    south_coordinates_file.write('id,longitude,latitude\n')
    north_coord = []
    south_coord = []

    mon_json = {
        "type": "FeatureCollection",
        "features": []
    }
    print("generating the files...")
    for i in range(sheet_obj.max_row):
        north = sheet_obj.cell(row=i + 1, column=8)
        south = sheet_obj.cell(row=i + 1, column=9)
        if (north.value is not None or south.value is not None):
            if(not(north.value=='North coordinates')):
                north_coordinates_file.write(str(i) + ',' + north.value + '\n')
                north_coord.append(float(str(north.value).split(', ')[1]))
                north_coord.append(float(str(north.value).split(', ')[0]))
            if (not (south.value == 'South coordinates')):
                south_coordinates_file.write(str(i) + ',' + south.value + '\n')
                south_coord.append(float(str(south.value).split(', ')[1]))
                south_coord.append(float(str(south.value).split(', ')[0]))
                tmp = {
                    "type": "Feature",
                    "properties": {},
                    "geometry": {
                        "coordinates": [
                            north_coord,
                            south_coord
                        ],
                        "type": "LineString"
                    }
                }

                mon_json["features"].append(tmp)
                with open('files/geojson.json', 'w') as mon_fichier:

                    json.dump(mon_json, mon_fichier,indent=4)
                north_coord = []
                south_coord = []
    print("files generated in files directory")
    # Print value of cell object
    # using the value attribute
    #print(cell_obj.value)
# See PyCharm help at https://www.jetbrains.com/help/pycharm/