from math import *
import warnings
from collections import OrderedDict
import numpy as np
import openpyxl
import pandas
import statistics as stat
import matplotlib.pyplot as plt
import pylab
import pyproj
from numpy import polyfit
from scipy.stats import linregress
from scipy.optimize import curve_fit
import io
import numpy as np
from jinja2 import Environment, PackageLoader, select_autoescape, FileSystemLoader


def deg2rad(dd):
    return dd / 180 * pi
def distanceGPS(latA, longA, latB, longB):
    RT = 6378137
    S = acos(sin(latA) * sin(latB) + cos(latA) * cos(latB) * cos(abs(longB - longA)))
    return S * RT

def mean(min, max, d1):
    mean_result = 0
    sum =0
    angles = d1.keys()
    for angle in angles:
        if angle > min and angle < max:
            sum = sum + 1
            mean_result = mean_result + d1[angle]
    if sum !=0:
        mean_result = mean_result/sum
    else:
        mean_result = 0
    return mean_result

#trouve le P du dictionnaire en paramètres
def median(min, max, d1,p):
    angles = d1.keys()
    list_median = []
    for angle in angles:
        if angle > min and angle < max:
            list_median.append(d1[angle])
    # list_median = sorted(list_median)
    return np.percentile(np.array(list_median),p)

#détermine le P d'une liste

def update_dict(step, d1,p):
    angles = d1.keys() # get angles
    angles = sorted(angles)
    d2 = dict()
    for angle in angles:
        angle_value = round(angle)
        if angle_value%step == 0:
           #d2.update({ str(angle_value-step) + "<->" + str(angle_value) : median(angle_value - step, angle_value, d1)})
           d2.update({ str(angle_value) : median(angle_value - step, angle_value, d1,p)})
    return d2
def get_type(path) -> str:
    warnings.filterwarnings("ignore")
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    for i in range(sheet_obj.max_row):
        cell_obj = sheet_obj.cell(row=i + 1, column=1).value
        if cell_obj == 'coordinates_system':
            return sheet_obj.cell(row=i + 1, column=2).value

def isNetwork(string) :
    if string is not None:
        mc = str(string).find("MC")
        gw = str(string).find("GW")
        return gw != -1 and mc != -1

def pos_TC(TC,path) -> str: #TC1.1.1
    tc = TC.replace('TC','')
    l = tc.split('.')
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    for i in range(sheet_obj.max_row):
        cell_obj = sheet_obj.cell(row=i + 1, column=10)
        if cell_obj.value != "TC coordinates" and cell_obj.value is not None:
            if sheet_obj.cell(row=i + 1, column=1).value == int(l[0]) and sheet_obj.cell(row=i + 1, column=2).value == int(l[1]) and sheet_obj.cell(row=i + 1, column=3).value == int(l[2]):
                return sheet_obj.cell(row=i + 1, column=10).value
    return 0

def pos_GW(TC,path) -> str:
    tc = TC.replace('TC', '')
    l = tc.split('.')
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    for i in range(sheet_obj.max_row):
        cell_obj = sheet_obj.cell(row=i + 1, column=1)
        if isNetwork(cell_obj.value) and str(cell_obj.value).replace('GW','').replace('MC','').split('_')[1] == l[1]:
            return str(sheet_obj.cell(row=i + 1, column=5).value)

def transform(x,y) -> list:
    inProj = pyproj.Proj(init='epsg:2154')
    outProj = pyproj.Proj(init='epsg:4326')
    x2, y2 = pyproj.transform(inProj, outProj, x, y)
    return [x2, y2]

def dist(TC,path) -> float: #retourne la distance d'un TC à sa GW en mètres
    coord_GW = pos_GW(TC,path).replace(' ' , '').split(',')
    pos = pos_TC(TC,path)
    coord_TC = str(pos_TC(TC,path)).replace(' ' , '').split(',')
    if get_type(path) != 4326:
        coord_GW = transform(coord_GW[0],coord_GW[1])
        coord_TC = transform(coord_TC[0], coord_TC[1])
    lon_GW = deg2rad(float(coord_GW[1]))
    lat_GW = deg2rad(float(coord_GW[0]))
    lon_TC = deg2rad(float(coord_TC[1]))
    lat_TC = deg2rad(float(coord_TC[0]))
    return distanceGPS(lat_TC,lon_TC,lat_GW,lon_GW)
def get_unique_trackers():
    df = pandas.read_csv('data_aws.csv')
    trackers = df["tracker"].tolist()
    unique_trackers = set(trackers)
    unique_trackers = sorted(unique_trackers)
    return unique_trackers
def create_dict(keys,values) -> dict:
    d = dict()
    for i in range(len(keys)):
        d.update({keys[i] : values[i]})
    return d
def countnan(l):
    count = 0
    for x in range(len(l)):  # suppression des nan
        val = l[x]
        if isnan(val) == True:
            count = count + 1
    return count
def nan_distance(Y):
    step = 15
    print("making distance latence ...")
    df = pandas.read_csv('data_aws.csv')
    trackers_distance = create_dict(get_unique_trackers(), getdist("config_sernhac.xlsx"))
    trackers_distance = dict(sorted(trackers_distance.items(), key=lambda item: item[1]))
    unique_trackers = trackers_distance.keys()
    nan_dist = {"nbNan": list(), "dist": trackers_distance.keys()}
    for tc in unique_trackers:
        tc_data = df.loc[df['tracker'] == tc]
        Ylatence = tc_data[Y].tolist()
        nan_dist["nbNan"].append(countnan(Ylatence))
    dist = nan_dist["dist"] # get angles
    d2 = dict()
    for d in dist:
        dist_value = round(d)
        if dist_value%step == 0:
           d2.update({ str(dist_value) : median(dist_value - step, dist_value, nan_dist,50)})
    print(nan_dist)
    plt.plot(nan_dist["dist"], nan_dist["nbNan"])
    plt.ylabel('nb_nan')
    if Y == 'zigbee_rx_signal_strength':
        plt.savefig('template/lat/courbe_nandist_rec.png')
        print("courbe_nandist_rec drawed !")
    else:
        plt.savefig('template/lat/courbe_nandist_lat.png')
        print("courbe_nandist_lat drawed !")


def draw_curves(X, Y):
    print("parsing csv file ...")
    df = pandas.read_csv('data_aws.csv')
    trackers_distance = create_dict(get_unique_trackers(),getdist("config_sernhac.xlsx"))
    trackers_distance = dict(sorted(trackers_distance.items(), key=lambda item: item[1]))
    unique_trackers = trackers_distance.keys()
    print("csv parsed !")
    images = []
    for tc in list(unique_trackers):
        print(tc)
        step = 15
        dmed = dict()
        dproved =dict()
        dpossible = dict()
        tc_data = df.loc[df['tracker'] == tc]
        Xcurrent_angles = tc_data[X].tolist()
        Ylatence = tc_data[Y].tolist()
        Ylatence2 = list()
        Xcurrent_angles2 = list()
        for x in range(len(Ylatence)):#suppression des nan
            val = Ylatence[x]
            if isnan(val) == False:
                Ylatence2.append(val)
                Xcurrent_angles2.append(Xcurrent_angles[x])
        for i in range(len(Xcurrent_angles2)):
            dmed.update({Xcurrent_angles2[i] : Ylatence2[i]}) # {angle : latence}
            dproved.update({Xcurrent_angles2[i]: Ylatence2[i]})  # {angle : latence}
            dpossible.update({Xcurrent_angles2[i]: Ylatence2[i]})  # {angle : latence}
        dmed = update_dict(step, dmed,50)
        dproved = update_dict(step, dproved,90)
        dpossible = update_dict(step, dpossible, 10)
        x = list(dmed.keys())
        plt.plot(x,dmed.values(),x,dproved.values(),x, dpossible.values())
        plt.title(
            "Distance TC <-> GW : "+str(round(dist(tc,"config_sernhac.xlsx")))+" m" +
            "\n nombre de nan : " + str(countnan(Ylatence))
        )
        print("Distance TC <-> GW : "+str(round(dist(tc,"config_sernhac.xlsx")))+" m")
        print("nombre de nan : " + str(countnan(Ylatence)))
        plt.gca().legend(('median','P90','P10'))
        plt.xlabel('angle')
        src = 'courbe_' + tc + '.png'
        if Y == 'zigbee_rx_signal_strength':
            plt.ylabel('reception')
            plt.ylim([0,120])
            plt.savefig('template/rec/courbe_' + tc + '.png')
        else:
            plt.ylabel('latence')
            plt.ylim([0, 1])
            plt.savefig('template/lat/courbe_' + tc + '.png')
        plt.show()
        images.append(str(src))
    return images

def draw_all_curves(path):
    Y = ['zigbee_last_message_txrx', 'zigbee_rx_signal_strength']
    X = 'current_angle'
    lat_images = []
    rec_images = []
    all = dict()
    for y in Y:
#        all.update(draw_curves(X, y))
        if y == 'zigbee_last_message_txrx':
            print("latence step")
            lat_images = draw_curves(X, y)
            #nan_distance(y)
        else:
            print("reception step")
            rec_images = draw_curves(X, y)
            #nan_distance(y)
    all_img = lat_images + rec_images
    #all = {"img" : all_img,"dist" : getdist(path)}
    return all_img
def getdist(path):
    l = []
    df = pandas.read_csv('data_aws.csv')
    trackers = df["tracker"].tolist()
    unique_trackers = set(trackers)
    unique_trackers = sorted(unique_trackers)
    for tc in unique_trackers:
        l.append(round(dist(tc,path)))
    return l
#renvoie la dictionnaire avec distance et latence
def lat_list_med(path) -> dict:
    d1 = dict()
    d2 = dict()
    l = list()
    print("parsing csv file ...")
    df = pandas.read_csv('data_aws.csv')
    unique_trackers = get_unique_trackers()
    for tc in list(unique_trackers):
        GW = tc.split('.')[1]
        if GW == str(1):
            tc_data = df.loc[df['tracker'] == tc]
            lat = tc_data['zigbee_last_message_txrx'].tolist()
            for val in lat:#suppression des nan
                if isnan(val) == False:
                    l.append(val)
            d1.update({dist(tc,path) : np.percentile(np.array(l), 50)})
            d1 = dict(OrderedDict(sorted(d1.items())))
            #d = update_dict(100, d,50)
            plt.scatter(d1.keys(),d1.values(), color = 'green')

        if GW == str(2):
            tc_data = df.loc[df['tracker'] == tc]
            lat = tc_data['zigbee_last_message_txrx'].tolist()
            for val in lat:#suppression des nan
                if isnan(val) == False:
                    l.append(val)
            d2.update({dist(tc,path) : np.percentile(np.array(l), 50)})
            d2 = dict(OrderedDict(sorted(d2.items())))
            #d = update_dict(100, d,50)
            plt.scatter(d2.keys(),d2.values(), color = 'blue')
    plt.ylabel('latence')
    plt.xlabel('distance')
    plt.show()
    return d1


# env = Environment(loader = FileSystemLoader("template"))
# template = env.get_template("mytemplate.html.j2")
# output = template.render(images = draw_all_curves("config_sernhac.xlsx"))
# with io.open("index3.html", "w") as file_point:
#     file_point.write(output)
print(get_unique_trackers())
print(lat_list_med("config_sernhac.xlsx"))
# print("TC: " + str(pos_TC("TC1.1.1", "config_sernhac.xlsx")))
# print("GW: " + str(pos_GW("TC1.1.1", "config_sernhac.xlsx")))
#print("dist : " + str(round(dist("TC1.2.84", "config_gargenville.xlsx"))))
#print(getdist("config_sernhac.xlsx"))
#print(draw_all_curves("config_sernhac.xlsx"))
#nan_distance('zigbee_last_message_txrx')
###############################################################
#geopy